import openpyxl, json, datetime

SRC = '/mnt/user-data/uploads/Danagram_4_Letter_Word_List.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['Sheet1']

# canonicalize the drifted note vocabulary
CANON = {
 'subtlex only':'SUBTLEX Only','proper':'Proper','propper':'Proper','prof':'Proper','proper?':'Proper',
 'slang':'Slang','foreign':'Foreign','profane':'Profane','old':'Old','odd':'Odd','jargon':'Jargon',
 '?':'?','exclamation':'Exclamation','contraction':'Contraction','onomatopoeia':'Onomatopoeia',
 'shortening':'Shortening','prefix':'Prefix','just for fun!':'?',
}
def canon(n):
    if n is None: return None
    k = str(n).strip().lower()
    if k == '': return None
    return CANON.get(k, str(n).strip())  # unknown notes preserved as-is

corpus = []      # static, frequency-ordered
progress = {}    # word -> {check, gen, note}  (only deliberately-decided words)
undecided = 0

for r in range(2, ws.max_row+1):
    note = canon(ws.cell(r,2).value)
    check = bool(ws.cell(r,3).value)
    gen   = bool(ws.cell(r,4).value)
    word  = ws.cell(r,5).value
    if not word: continue
    en  = 1 if ws.cell(r,6).value else 0
    sub = 1 if ws.cell(r,7).value else 0
    scr = 1 if ws.cell(r,8).value else 0
    freq = ws.cell(r,9).value or 0
    fpm  = ws.cell(r,13).value or 0   # SUBTLWF, freq per million
    corpus.append({'word':word,'enable1':en,'subtlex':sub,'scrabble':scr,
                   'freq':int(freq),'fpm':round(float(fpm),2)})
    # "untouched" = both false AND no note -> stays in the swipe queue
    if not check and not gen and note is None:
        undecided += 1
    else:
        progress[word] = {'check':check,'gen':gen,'note':note}

stamp = datetime.datetime.utcnow().isoformat()+'Z'
meta = {'generated':stamp,'total':len(corpus),'decided':len(progress),'queue':undecided}

import os
os.makedirs('site', exist_ok=True)
json.dump(corpus, open('site/words.json','w'), separators=(',',':'))
json.dump(progress, open('site/progress.json','w'), indent=0)
json.dump({}, open('site/definitions.json','w'))
json.dump(meta, open('site/meta.json','w'), indent=2)

print(json.dumps(meta, indent=2))
print('words.json bytes:', os.path.getsize('site/words.json'))
print('progress.json bytes:', os.path.getsize('site/progress.json'))
