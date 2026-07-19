#!/usr/bin/env python3
"""Recombine the JSON source of truth back into a .xlsx matching the original layout.
Usage: python rebuild_xlsx.py [site_dir] [out.xlsx]"""
import json, sys, openpyxl
from openpyxl.styles import Font

site = sys.argv[1] if len(sys.argv)>1 else 'site'
out  = sys.argv[2] if len(sys.argv)>2 else 'Danagram_4_Letter_Word_List_rebuilt.xlsx'

corpus   = json.load(open(f'{site}/words.json'))
progress = json.load(open(f'{site}/progress.json'))

wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Sheet1'
headers = ['','Notes','Check','Gen','word','enable1','subtlex','scrabble','FREQcount','SUBTLWF']
ws.append(headers)
for c in ws[1]: c.font = Font(name='Arial', bold=True)

for rec in corpus:                      # preserves frequency order
    w = rec['word']
    d = progress.get(w, {'check':False,'gen':False,'note':None})
    ws.append(['', d.get('note'), bool(d['check']), bool(d['gen']), w,
               bool(rec['enable1']), bool(rec['subtlex']), bool(rec['scrabble']),
               rec['freq'], rec['fpm']])
for row in ws.iter_rows(min_row=2):
    for c in row: c.font = Font(name='Arial')

wb.save(out)
print(f'wrote {out}  ({len(corpus)} rows, {len(progress)} decisions applied)')
